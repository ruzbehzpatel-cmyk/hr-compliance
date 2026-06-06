"""
generate_dashboard.py — builds dashboard.html from HR_Compliance_Framework_COMPLETE.xlsx

Sections rendered:
  1. Interactive US state map (click a state to see all its regs grouped by domain, with CoE owner)
  2. Today's changes (most recent Updates YYYY-MM-DD tab)
  3. KPI cards
  4. Upcoming effective dates + Top jurisdictions
  5. Domain donut + Recent changes feed
  6. Filterable regulation table
"""
import openpyxl
import json
import re
import argparse
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

LAW_TABS = ["Federal Laws", "State Laws", "Additional & Emerging"]
KEY_FIELDS = [
    "Law ID", "Law / Regulation Name", "Jurisdiction", "Core Domain",
    "Sub-Domain(s)", "Category", "Summary / Key Requirements", "Applicability",
    "Effective Date", "Status", "Gov Source URL", "Risk Level",
    "Penalty Exposure", "Workforce %", "Primary HR Owner", "Compliance Control",
    "Last Review", "Current Gap", "Change Watch", "Notes / Action Items",
]

# Full state name lookup (matches us-atlas TopoJSON `properties.name` exactly)
US_STATES = {
    "Alabama","Alaska","Arizona","Arkansas","California","Colorado","Connecticut",
    "Delaware","District of Columbia","Florida","Georgia","Hawaii","Idaho","Illinois",
    "Indiana","Iowa","Kansas","Kentucky","Louisiana","Maine","Maryland","Massachusetts",
    "Michigan","Minnesota","Mississippi","Missouri","Montana","Nebraska","Nevada",
    "New Hampshire","New Jersey","New Mexico","New York","North Carolina","North Dakota",
    "Ohio","Oklahoma","Oregon","Pennsylvania","Rhode Island","South Carolina",
    "South Dakota","Tennessee","Texas","Utah","Vermont","Virginia","Washington",
    "West Virginia","Wisconsin","Wyoming",
}

# ---------- Data extraction ----------

def load_all_tab_rows(xlsx_path):
    """Load rows from law tabs AND survey tabs. Returns (law_rows, all_rows_for_state_view, update_tabs)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    law_rows = []
    all_rows = []  # includes survey tab rows for state view
    update_tabs = []
    for sheet_name in wb.sheetnames:
        if re.match(r"Updates? \d{4}-\d{2}-\d{2}", sheet_name):
            update_tabs.append(sheet_name)
            continue
        if sheet_name in ("Reference", "Dashboard", "QC Findings"):
            continue
        is_law_tab = sheet_name in LAW_TABS
        is_survey_tab = sheet_name.startswith("Survey")
        if not (is_law_tab or is_survey_tab):
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers) if h}
        for r in range(2, ws.max_row + 1):
            raw = [ws.cell(r, c + 1).value for c in range(len(headers))]
            if not any(v not in (None, "") for v in raw):
                continue
            row = {"__tab__": sheet_name}
            for fld in KEY_FIELDS:
                if fld in idx:
                    val = raw[idx[fld]]
                    row[fld] = "" if val is None else str(val)
                else:
                    row[fld] = ""
            all_rows.append(row)
            if is_law_tab:
                law_rows.append(row)
    return law_rows, all_rows, sorted(update_tabs)


def load_latest_updates(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    update_tabs = sorted([s for s in wb.sheetnames if re.match(r"Updates? \d{4}-\d{2}-\d{2}", s)])
    if not update_tabs:
        return {"tab_name": None, "tab_date": None, "rows": []}
    latest = update_tabs[-1]
    m = re.search(r"(\d{4}-\d{2}-\d{2})", latest)
    tab_date = m.group(1) if m else None
    ws = wb[latest]
    headers = [str(c.value or "").strip().lower() for c in ws[1]]
    def find(pred):
        for i, h in enumerate(headers):
            if pred(h):
                return i
        return -1
    cols = {
        "type":         find(lambda h: h.startswith("type")),
        "affected_tab": find(lambda h: "affected" in h),
        "law_id":       find(lambda h: "law id" in h),
        "name":         find(lambda h: "regulation name" in h or h == "law / regulation name" or ("law" in h and "name" in h)),
        "jurisdiction": find(lambda h: "jurisdiction" in h),
        "field":        find(lambda h: "field" in h),
        "before":       find(lambda h: h == "before"),
        "after":        find(lambda h: h == "after"),
        "url":          find(lambda h: "source" in h and "url" in h),
        "confidence":   find(lambda h: "confidence" in h),
        "notes":        find(lambda h: h.startswith("notes")),
    }
    rows = []
    for r in range(2, ws.max_row + 1):
        raw = [ws.cell(r, c + 1).value for c in range(len(headers))]
        if not any(v not in (None, "") for v in raw):
            continue
        def cell(key):
            i = cols.get(key, -1)
            return "" if i < 0 or raw[i] is None else str(raw[i])
        rows.append({
            "type": cell("type"), "affected_tab": cell("affected_tab"),
            "law_id": cell("law_id"), "name": cell("name"),
            "jurisdiction": cell("jurisdiction"), "field": cell("field"),
            "before": cell("before"), "after": cell("after"),
            "url": cell("url"), "confidence": cell("confidence"), "notes": cell("notes"),
        })
    return {"tab_name": latest, "tab_date": tab_date, "rows": rows}

# ---------- Derived metrics ----------

DATE_PATTERNS = [
    (re.compile(r"(\d{1,2})/(\d{1,2})/(\d{4})"), lambda m: date(int(m.group(3)), int(m.group(1)), int(m.group(2)))),
    (re.compile(r"(\d{4})-(\d{2})-(\d{2})"), lambda m: date(int(m.group(1)), int(m.group(2)), int(m.group(3)))),
]

def extract_first_date(text):
    if not text: return None
    for pattern, ctor in DATE_PATTERNS:
        for m in pattern.finditer(text):
            try: return ctor(m)
            except (ValueError, AttributeError): continue
    return None

def extract_future_date(text, today):
    if not text: return None
    candidates = []
    for pattern, ctor in DATE_PATTERNS:
        for m in pattern.finditer(text):
            try:
                d = ctor(m)
                if d >= today: candidates.append(d)
            except (ValueError, AttributeError): continue
    return min(candidates) if candidates else None

def normalize_risk(risk_str):
    s = (risk_str or "").upper()
    if "RED" in s: return "RED"
    if "YELLOW" in s: return "YELLOW"
    if "GREEN" in s: return "GREEN"
    return "UNKNOWN"

def normalize_category(c):
    """Strip sub-flavor suffixes so 'Pay Transparency – Proposed' → 'Pay Transparency'."""
    if not c: return ""
    c = str(c).strip()
    # Split on en dash, em dash, hyphen, or parenthesis
    for sep in [' – ', ' — ', ' - ', ' (']:
        if sep in c:
            c = c.split(sep)[0].strip()
            break
    return c

def normalize_jurisdiction(j):
    """Map raw jurisdiction string to a clean state name (or 'Federal' / None)."""
    if not j: return None
    j = j.strip()
    if j in US_STATES: return j
    if j.lower() == "federal" or "federal" in j.lower(): return "Federal"
    # Try matching state name within free-text jurisdictions
    for s in US_STATES:
        if s.lower() == j.lower(): return s
    return None

def compute_state_data(all_rows):
    """Group regs by state. Dedupe by Law ID across tabs. Returns {state: {count, by_domain: {domain: [reg, ...]}}}."""
    # Dedupe by Law ID (BOTH-TABS rule means many regs appear in 2+ tabs)
    by_law_id = {}
    for r in all_rows:
        lid = r.get("Law ID")
        if not lid: continue
        if lid not in by_law_id:
            by_law_id[lid] = {**r, "appears_in": [r["__tab__"]]}
        else:
            existing = by_law_id[lid]
            if r["__tab__"] not in existing["appears_in"]:
                existing["appears_in"].append(r["__tab__"])
    deduped = list(by_law_id.values())

    state_data = defaultdict(lambda: {"count": 0, "by_domain": defaultdict(list)})
    for r in deduped:
        state = normalize_jurisdiction(r.get("Jurisdiction"))
        if not state: continue
        domain = r.get("Core Domain") or "Other"
        state_data[state]["count"] += 1
        raw_cat = (r.get("Category") or "").strip()
        cat_norm = normalize_category(raw_cat)
        state_data[state]["by_domain"][domain].append({
            "law_id": r["Law ID"],
            "name": r["Law / Regulation Name"],
            "owner": r["Primary HR Owner"] or "—",
            "status": r["Status"],
            "effective": r["Effective Date"],
            "risk": normalize_risk(r["Risk Level"]),
            "url": r["Gov Source URL"],
            "appears_in": r["appears_in"],
            "category": raw_cat,
            "category_norm": cat_norm,
        })
    # Convert defaultdicts → dicts for JSON
    out = {}
    for state, d in state_data.items():
        out[state] = {
            "count": d["count"],
            "by_domain": {k: sorted(v, key=lambda x: x["name"]) for k, v in d["by_domain"].items()},
        }
    return out

def collect_categories(state_data):
    """Return sorted list of unique normalized categories, with a count per category."""
    from collections import Counter
    cnt = Counter()
    for state, d in state_data.items():
        for domain, regs in d["by_domain"].items():
            for r in regs:
                cn = r.get("category_norm", "")
                if cn:
                    cnt[cn] += 1
    return [{"name": k, "count": v} for k, v in sorted(cnt.items(), key=lambda x: (-x[1], x[0]))]

def compute_metrics(rows):
    today = date.today()
    horizon = date(today.year + 1, today.month, today.day) if not (today.month == 2 and today.day == 29) else date(today.year + 1, 3, 1)
    metrics = {
        "total": len(rows),
        "by_tab": dict(Counter(r["__tab__"] for r in rows)),
        "by_risk": dict(Counter(normalize_risk(r.get("Risk Level")) for r in rows)),
        "by_jurisdiction": dict(Counter(r["Jurisdiction"] for r in rows if r["Jurisdiction"]).most_common(10)),
        "by_domain": dict(Counter(r["Core Domain"] for r in rows if r["Core Domain"]).most_common(10)),
        "effective_next_90": [], "effective_next_12mo": [], "recent_changes": [],
        "high_risk_count": 0,
    }
    ninety_days = date.fromordinal(today.toordinal() + 90)
    thirty_days_ago = date.fromordinal(today.toordinal() - 30)
    for r in rows:
        if normalize_risk(r.get("Risk Level")) == "RED":
            metrics["high_risk_count"] += 1
        nxt = extract_future_date(r.get("Effective Date", ""), today)
        if nxt:
            entry = {
                "law_id": r["Law ID"], "name": r["Law / Regulation Name"],
                "jurisdiction": r["Jurisdiction"], "domain": r["Core Domain"],
                "risk": normalize_risk(r.get("Risk Level")), "effective": nxt.isoformat(),
                "effective_raw": r["Effective Date"], "tab": r["__tab__"], "url": r["Gov Source URL"],
            }
            if nxt <= horizon: metrics["effective_next_12mo"].append(entry)
            if nxt <= ninety_days: metrics["effective_next_90"].append(entry)
        lr_date = extract_first_date(r.get("Last Review", ""))
        if lr_date and lr_date >= thirty_days_ago:
            metrics["recent_changes"].append({
                "law_id": r["Law ID"], "name": r["Law / Regulation Name"],
                "jurisdiction": r["Jurisdiction"], "tab": r["__tab__"],
                "status": r["Status"], "last_review": lr_date.isoformat(),
            })
    def dedupe(entries, key, reverse=False):
        seen = {}
        for e in entries:
            lid = e.get("law_id") or ""
            if not lid: continue
            if lid not in seen:
                seen[lid] = e
            else:
                better = (e[key] < seen[lid][key]) if not reverse else (e[key] > seen[lid][key])
                if better: seen[lid] = e
        return list(seen.values())
    metrics["effective_next_12mo"] = sorted(dedupe(metrics["effective_next_12mo"], "effective"), key=lambda x: x["effective"])
    metrics["effective_next_90"]   = sorted(dedupe(metrics["effective_next_90"], "effective"), key=lambda x: x["effective"])
    metrics["recent_changes"]      = sorted(dedupe(metrics["recent_changes"], "last_review", reverse=True), key=lambda x: x["last_review"], reverse=True)
    metrics["effective_next_90_count"] = len(metrics["effective_next_90"])
    metrics["recent_changes_count"]    = len(metrics["recent_changes"])
    return metrics

# ---------- HTML render ----------

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>HR Compliance Dashboard — __TODAY__</title>
<style>
:root { --bg:#fff; --bg2:#f7f6f2; --bg3:#efede5; --text:#1a1a1a; --text2:#5a5a5a; --text3:#888;
  --border:rgba(0,0,0,.1); --border2:rgba(0,0,0,.18);
  --red-bg:#FCEBEB; --red-fg:#A32D2D; --yellow-bg:#FAEEDA; --yellow-fg:#854F0B;
  --green-bg:#EAF3DE; --green-fg:#3B6D11; --blue-bg:#E6F1FB; --blue-fg:#0C447C;
  --purple-bg:#EEEDFE; --purple-fg:#3C3489; --coral-bg:#FAECE7; --coral-fg:#712B13;
  --teal-bg:#E1F5EE; --teal-fg:#085041;
  --purple:#534AB7; --teal:#1D9E75; --coral:#D85A30; --gray:#888; }
@media (prefers-color-scheme: dark) { :root { --bg:#1a1a1a; --bg2:#222; --bg3:#2a2a2a; --text:#f0f0f0; --text2:#bbb; --text3:#888; --border:rgba(255,255,255,.1); --border2:rgba(255,255,255,.2); } }
* { box-sizing: border-box; }
body { margin:0; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif; background:var(--bg3); color:var(--text); line-height:1.5; }
.app { max-width:1400px; margin:0 auto; padding:24px; }
.refresh-banner { background:var(--bg2); border:0.5px solid var(--border); border-radius:8px; padding:10px 14px; margin-bottom:14px; font-size:12px; }
.refresh-banner-row { display:flex; flex-wrap:wrap; gap:6px; align-items:center; }
.refresh-label { color:var(--text3); text-transform:uppercase; letter-spacing:0.5px; font-size:10px; font-weight:500; }
.refresh-value { color:var(--text); font-weight:500; font-family:Menlo,Monaco,'SF Mono',monospace; font-size:11px; }
.refresh-sep { color:var(--text3); opacity:0.5; margin:0 4px; }
.header { display:flex; justify-content:space-between; align-items:baseline; margin-bottom:16px; flex-wrap:wrap; gap:12px; }
.title { font-size:22px; font-weight:500; }
.subtitle { font-size:12px; color:var(--text2); }
.map-card { background:var(--bg); border:0.5px solid var(--border); border-radius:12px; padding:16px; margin-bottom:16px; }
.map-head { display:flex; justify-content:space-between; align-items:center; margin-bottom:12px; flex-wrap:wrap; gap:8px; }
.map-actions { display:flex; gap:8px; }
.map-actions button { font:inherit; font-size:12px; padding:6px 12px; border:0.5px solid var(--border2); border-radius:6px; background:var(--bg); color:var(--text); cursor:pointer; }
.map-actions button:hover { background:var(--bg2); }
.map-actions button.primary { background:var(--purple); color:#fff; border-color:var(--purple); }
.map-wrap { position:relative; width:100%; max-width:900px; margin:0 auto; padding:8px 0; }
.map-wrap svg { width:100%; height:auto; display:block; overflow:visible; }
.us-outline { fill:#efede5; stroke:#bbb; stroke-width:1; }
@media (prefers-color-scheme: dark) { .us-outline { fill:#2a2a2a; stroke:#555; } }
.callout-line { stroke:#888; stroke-width:1; fill:none; opacity:0.6; }
.pill-group { cursor:pointer; }
.pill-group .pill-bg { fill:transparent; }
.pill-group:hover .pill-bg { fill:rgba(83,74,183,0.12); }
.pill-abbr { font-size:9px; font-weight:500; fill:#5a5a5a; text-anchor:middle; pointer-events:none; }
.pill-count { font-size:14px; font-weight:500; fill:#26215C; text-anchor:middle; pointer-events:none; }
.pill-count.tier-3 { fill:#26215C; }
.pill-count.tier-2 { fill:#3C3489; }
.pill-count.tier-1 { fill:#534AB7; }
.pill-count.tier-0 { fill:#888; font-size:11px; font-weight:400; }
@media (prefers-color-scheme: dark) {
  .pill-abbr { fill:#bbb; }
  .pill-count { fill:#CECBF6; }
  .pill-count.tier-3 { fill:#fff; }
  .pill-count.tier-2 { fill:#CECBF6; }
  .pill-count.tier-1 { fill:#AFA9EC; }
}
.map-legend { display:flex; align-items:center; gap:6px; font-size:11px; color:var(--text2); margin-top:8px; justify-content:center; }
.map-legend .swatch { display:inline-flex; align-items:center; gap:4px; padding:2px 6px; border-radius:8px; }
.map-legend .swatch.t3 { background:rgba(38,33,92,0.12); color:#26215C; }
.map-legend .swatch.t2 { background:rgba(60,52,137,0.10); color:#3C3489; }
.map-legend .swatch.t1 { background:rgba(83,74,183,0.10); color:#534AB7; }
.map-legend .swatch.t0 { color:#888; }
.state-panel { background:var(--bg2); border-radius:8px; padding:16px; margin-top:16px; display:none; }
.state-panel.open { display:block; }
.state-panel-head { display:flex; justify-content:space-between; align-items:center; margin-bottom:12px; }
.state-panel-title { font-size:16px; font-weight:500; }
.state-panel-close { background:none; border:none; color:var(--text2); font-size:18px; cursor:pointer; padding:4px 8px; }
.state-panel-close:hover { color:var(--text); }
.domain-group { margin-bottom:14px; }
.domain-group-head { font-size:12px; font-weight:500; color:var(--text2); text-transform:uppercase; letter-spacing:0.5px; padding:6px 0; border-bottom:0.5px solid var(--border); margin-bottom:6px; }
.domain-reg { display:flex; gap:10px; align-items:flex-start; padding:8px 0; border-bottom:0.5px solid var(--border); }
.domain-reg:last-child { border-bottom:none; }
.domain-reg-name { flex:1; font-size:13px; }
.domain-reg-id { color:var(--text3); font-size:10px; margin-left:4px; }
.domain-reg-meta { font-size:11px; color:var(--text3); margin-top:2px; }
.coe { display:inline-block; padding:3px 10px; border-radius:12px; font-size:11px; background:var(--teal-bg); color:var(--teal-fg); white-space:nowrap; }
.coe.unowned { background:var(--bg3); color:var(--text3); }
.today-card { background:var(--bg); border:0.5px solid var(--border); border-left:4px solid var(--coral); border-radius:12px; padding:16px; margin-bottom:16px; }
.today-card.empty { border-left-color:var(--gray); }
.today-card.stale { border-left-color:var(--yellow-fg); background:var(--yellow-bg); }
.today-head { display:flex; justify-content:space-between; align-items:center; margin-bottom:8px; flex-wrap:wrap; gap:8px; }
.today-title { font-size:15px; font-weight:500; display:flex; align-items:center; gap:8px; }
.today-badges { display:flex; gap:6px; flex-wrap:wrap; }
.tag { display:inline-block; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:500; white-space:nowrap; }
.tag.CHANGE { background:var(--blue-bg); color:var(--blue-fg); }
.tag.NEW { background:var(--green-bg); color:var(--green-fg); }
.tag.SPLIT { background:var(--purple-bg); color:var(--purple-fg); }
.tag.CATCH-UP, .tag.CATCHUP { background:var(--coral-bg); color:var(--coral-fg); }
.tag.OBSERVATION { background:var(--bg3); color:var(--text3); }
.tag.UNKNOWN { background:var(--bg3); color:var(--text3); }
.tag.HIGH { background:var(--red-bg); color:var(--red-fg); }
.tag.MED { background:var(--yellow-bg); color:var(--yellow-fg); }
.tag.LOW { background:var(--green-bg); color:var(--green-fg); }
.today-empty { color:var(--text2); font-size:13px; padding:8px 0; }
.today-summary { padding:14px 16px; background:var(--bg2); border-left:3px solid var(--coral); border-radius:8px; margin-bottom:14px; font-size:13px; }
.today-summary-line { color:var(--text); margin-bottom:10px; font-size:14px; }
.today-summary-line strong { color:var(--text); font-weight:500; }
.today-highlights { font-size:12px; }
.today-highlights-label { font-size:11px; color:var(--text3); text-transform:uppercase; letter-spacing:0.5px; margin-bottom:6px; margin-top:4px; font-weight:500; }
.today-highlight { display:flex; gap:8px; align-items:flex-start; padding:6px 0; }
.today-highlight-tag { font-size:10px; padding:2px 6px; border-radius:8px; font-weight:500; flex-shrink:0; margin-top:1px; }
.today-highlight-text { color:var(--text); line-height:1.4; }
.today-highlight-text .frame { color:var(--text3); }
.today-list { display:flex; flex-direction:column; gap:0; }
.today-item { padding:10px 0; border-top:0.5px solid var(--border); cursor:pointer; }
.today-item:first-child { border-top:none; }
.today-item:hover { background:var(--bg2); margin:0 -16px; padding:10px 16px; }
.today-item-head { display:flex; gap:10px; align-items:center; flex-wrap:wrap; }
.today-item-name { font-size:13px; font-weight:500; flex:1; min-width:200px; }
.today-item-meta { font-size:11px; color:var(--text3); margin-top:2px; }
.today-item-detail { display:none; margin-top:8px; padding:8px 12px; background:var(--bg2); border-radius:6px; font-size:12px; }
.today-item.open .today-item-detail { display:block; }
.diff-row { margin:4px 0; }
.diff-label { color:var(--text3); font-size:10px; text-transform:uppercase; letter-spacing:0.5px; }
.diff-text { color:var(--text); white-space:pre-wrap; word-break:break-word; }
.kpis { display:grid; grid-template-columns:repeat(auto-fit, minmax(180px, 1fr)); gap:12px; margin-bottom:16px; }
.kpi { background:var(--bg2); border-radius:8px; padding:14px; }
.kpi-label { font-size:12px; color:var(--text2); }
.kpi-value { font-size:26px; font-weight:500; margin-top:2px; }
.kpi-sub { font-size:11px; color:var(--text3); margin-top:2px; }
.kpi.red { background:var(--red-bg); } .kpi.red .kpi-label, .kpi.red .kpi-value, .kpi.red .kpi-sub { color:var(--red-fg); }
.kpi.yellow { background:var(--yellow-bg); } .kpi.yellow .kpi-label, .kpi.yellow .kpi-value, .kpi.yellow .kpi-sub { color:var(--yellow-fg); }
.kpi.blue { background:var(--blue-bg); } .kpi.blue .kpi-label, .kpi.blue .kpi-value, .kpi.blue .kpi-sub { color:var(--blue-fg); }
.kpi.green { background:var(--green-bg); } .kpi.green .kpi-label, .kpi.green .kpi-value, .kpi.green .kpi-sub { color:var(--green-fg); }
.kpi.clickable { cursor:pointer; transition:transform .1s, box-shadow .1s; user-select:none; }
.kpi.clickable:hover { transform:translateY(-2px); box-shadow:0 4px 12px rgba(0,0,0,0.08); }
.kpi.clickable:focus { outline:2px solid var(--blue-fg); outline-offset:2px; }
.kpi.clickable.active { box-shadow:0 0 0 2px var(--text); }
.kpi-drill { display:none; background:var(--bg); border:0.5px solid var(--border); border-radius:12px; padding:16px; margin-bottom:16px; }
.kpi-drill.open { display:block; }
.kpi-drill-head { display:flex; justify-content:space-between; align-items:center; margin-bottom:12px; padding-bottom:10px; border-bottom:0.5px solid var(--border); }
.kpi-drill-title { font-size:15px; font-weight:500; }
.kpi-drill-close { background:none; border:none; color:var(--text2); font-size:20px; cursor:pointer; padding:4px 10px; }
.kpi-drill-close:hover { color:var(--text); }
.kpi-drill-body { max-height:480px; overflow-y:auto; }
.kpi-drill table { width:100%; font-size:12px; }
.kpi-drill th { text-align:left; color:var(--text2); padding:8px 6px; border-bottom:0.5px solid var(--border); font-weight:500; position:sticky; top:0; background:var(--bg); }
.kpi-drill td { padding:8px 6px; border-bottom:0.5px solid var(--border); vertical-align:top; }
.kpi-drill tr:hover td { background:var(--bg2); }
.kpi-drill .row-id { color:var(--text3); font-family:Menlo,Monaco,monospace; font-size:10px; white-space:nowrap; }
.kpi-drill .row-eff { color:var(--text2); font-size:11px; white-space:nowrap; }
.kpi-drill-empty { color:var(--text2); font-size:13px; padding:16px 0; }
.row { display:grid; gap:12px; margin-bottom:16px; }
.row.split-2 { grid-template-columns:1.5fr 1fr; }
.row.split-eq { grid-template-columns:1fr 1fr; }
@media (max-width:900px) { .row.split-2, .row.split-eq { grid-template-columns:1fr; } }
.card { background:var(--bg); border:0.5px solid var(--border); border-radius:12px; padding:16px; }
.card-head { display:flex; justify-content:space-between; align-items:center; margin-bottom:12px; }
.card-title { font-size:14px; font-weight:500; }
.card-link { font-size:12px; color:var(--blue-fg); text-decoration:none; cursor:pointer; }
table { width:100%; font-size:12px; border-collapse:collapse; }
th { text-align:left; color:var(--text2); padding:8px 6px; border-bottom:0.5px solid var(--border); font-weight:500; }
td { padding:8px 6px; border-bottom:0.5px solid var(--border); vertical-align:top; }
tr:hover td { background:var(--bg2); }
.pill { display:inline-block; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:500; }
.pill.RED { background:var(--red-bg); color:var(--red-fg); }
.pill.YELLOW { background:var(--yellow-bg); color:var(--yellow-fg); }
.pill.GREEN { background:var(--green-bg); color:var(--green-fg); }
.pill.UNKNOWN { background:var(--bg3); color:var(--text3); }
.feed-item { display:flex; gap:10px; align-items:flex-start; padding:8px 0; border-bottom:0.5px solid var(--border); font-size:12px; }
.feed-item:last-child { border-bottom:none; }
.feed-tag { background:var(--bg3); color:var(--text2); padding:2px 8px; border-radius:10px; font-size:10px; flex-shrink:0; }
.feed-meta { color:var(--text3); font-size:11px; }
.chart-wrap { position:relative; height:240px; }
.legend { display:flex; flex-wrap:wrap; gap:14px; font-size:11px; color:var(--text2); margin-bottom:8px; }
.legend-sw { display:inline-block; width:10px; height:10px; border-radius:2px; margin-right:4px; vertical-align:middle; }
.controls { display:flex; flex-wrap:wrap; gap:8px; align-items:center; margin-bottom:12px; }
.controls input, .controls select { font:inherit; font-size:12px; padding:6px 10px; border:0.5px solid var(--border2); border-radius:6px; background:var(--bg); color:var(--text); }
.controls input { flex:1; min-width:180px; }
.controls .count { font-size:11px; color:var(--text3); margin-left:auto; }
.reg-table { font-size:11px; }
.reg-table td { padding:6px; }
a { color:var(--blue-fg); }
.reg-tooltip { position:fixed; z-index:9999; max-width:420px; padding:10px 14px; background:#1a1a1a; color:#f0f0f0; border-radius:8px; box-shadow:0 8px 24px rgba(0,0,0,0.25); font-size:12px; line-height:1.5; pointer-events:none; opacity:0; transition:opacity .12s; }
.reg-tooltip.visible { opacity:1; }
.reg-tooltip .rt-title { font-size:13px; font-weight:500; margin-bottom:6px; color:#fff; }
.reg-tooltip .rt-meta { font-size:10px; color:#bbb; margin-bottom:6px; text-transform:uppercase; letter-spacing:0.5px; }
.reg-tooltip .rt-body { color:#e8e8e8; white-space:pre-wrap; word-wrap:break-word; }
.reg-tooltip .rt-empty { color:#888; font-style:italic; }
[data-summary] { cursor:help; }
[data-summary]:hover { background:var(--bg2); }
.kpi-drill tr[data-summary]:hover td, .reg-table tr[data-summary]:hover td { background:var(--bg2); }
.footer { font-size:11px; color:var(--text3); text-align:center; margin-top:24px; padding-top:16px; border-top:0.5px solid var(--border); }
</style>
</head>
<body>
<div class="app">

<div class="refresh-banner">
  <div class="refresh-banner-row">
    <span class="refresh-label">Last refresh</span>
    <span class="refresh-value">__GENERATED_AT__</span>
    <span class="refresh-sep">·</span>
    <span class="refresh-label">Source file</span>
    <span class="refresh-value" title="modified __SOURCE_MTIME__, __SOURCE_SIZE_KB__ KB">__SOURCE_FILE__</span>
    <span class="refresh-sep">·</span>
    <span class="refresh-label">File last modified</span>
    <span class="refresh-value">__SOURCE_MTIME__</span>
  </div>
</div>

<div class="header">
  <div>
    <div class="title">HR Compliance — live overview</div>
    <div class="subtitle">__TOTAL__ regulations tracked across Federal, State, and Additional &amp; Emerging tabs</div>
  </div>
</div>

<div class="kpis">
  <div class="kpi clickable" data-drill="all" tabindex="0" role="button" aria-label="Drill into all regulations"><div class="kpi-label">Total tracked</div><div class="kpi-value">__TOTAL__</div><div class="kpi-sub">__BY_TAB__</div></div>
  <div class="kpi red clickable" data-drill="red" tabindex="0" role="button" aria-label="Drill into RED risk regulations"><div class="kpi-label">RED risk</div><div class="kpi-value">__RED__</div><div class="kpi-sub">__RED_PCT__% of portfolio</div></div>
  <div class="kpi yellow clickable" data-drill="yellow" tabindex="0" role="button" aria-label="Drill into YELLOW risk regulations"><div class="kpi-label">YELLOW risk</div><div class="kpi-value">__YELLOW__</div><div class="kpi-sub">__YELLOW_PCT__% of portfolio</div></div>
  <div class="kpi blue clickable" data-drill="eff90" tabindex="0" role="button" aria-label="Drill into regulations effective in next 90 days"><div class="kpi-label">Effective ≤ 90 days</div><div class="kpi-value">__EFF90__</div><div class="kpi-sub">Action required</div></div>
  <div class="kpi green clickable" data-drill="recent" tabindex="0" role="button" aria-label="Drill into recent changes"><div class="kpi-label">Changes last 30 days</div><div class="kpi-value">__RECENT__</div><div class="kpi-sub">From daily scans</div></div>
</div>

<div class="kpi-drill" id="kpi-drill">
  <div class="kpi-drill-head">
    <div class="kpi-drill-title" id="kpi-drill-title"></div>
    <button class="kpi-drill-close" id="kpi-drill-close" aria-label="Close">×</button>
  </div>
  <div class="kpi-drill-body" id="kpi-drill-body"></div>
</div>

<div class="today-card" id="today-card">
  <div class="today-head">
    <div class="today-title" id="today-title">Today's changes</div>
    <div class="today-badges" id="today-badges"></div>
  </div>
  <div id="today-summary"></div>
  <div id="today-body"></div>
</div>

<div class="map-card">
  <div class="map-head">
    <div class="card-title">Regulations by state — click a state to see its regulations</div>
    <div class="map-actions">
      <label for="category-filter" style="font-size:12px;color:var(--text2);">Category:</label>
      <select id="category-filter" style="font-size:12px;padding:6px 10px;border:0.5px solid var(--border2);border-radius:6px;background:var(--bg);color:var(--text);min-width:200px;">
        <option value="">All categories (__TOTAL_REGS__)</option>
      </select>
      <button id="federal-btn">Federal (<span id="federal-count">0</span>)</button>
    </div>
  </div>
  <div class="map-wrap"><svg id="map-svg" viewBox="0 0 1000 600" preserveAspectRatio="xMidYMid meet" role="img" aria-label="US map with regulation counts per state"></svg></div>
  <div class="map-legend">
    <span>Regs per state:</span>
    <span class="swatch t1">1–low</span>
    <span class="swatch t2">medium</span>
    <span class="swatch t3">high</span>
    <span class="swatch t0">0 (none tracked)</span>
  </div>
  <div class="state-panel" id="state-panel">
    <div class="state-panel-head">
      <div class="state-panel-title" id="state-panel-title"></div>
      <button class="state-panel-close" id="state-panel-close" aria-label="Close">×</button>
    </div>
    <div id="state-panel-body"></div>
  </div>
</div>

<div class="row split-2">
  <div class="card">
    <div class="card-head"><div class="card-title">Upcoming effective dates — next 12 months</div><span class="card-link" id="upcoming-count"></span></div>
    <div style="max-height:380px; overflow-y:auto;">
      <table id="upcoming-table"><thead><tr><th>Date</th><th>Regulation</th><th>Juris.</th><th style="text-align:right;">Risk</th></tr></thead><tbody></tbody></table>
    </div>
  </div>
  <div class="card">
    <div class="card-head"><div class="card-title">Top jurisdictions</div></div>
    <div class="chart-wrap"><canvas id="jurisChart"></canvas></div>
  </div>
</div>

<div class="row split-eq">
  <div class="card">
    <div class="card-head"><div class="card-title">Portfolio by core domain</div></div>
    <div class="legend" id="domain-legend"></div>
    <div class="chart-wrap" style="height:220px;"><canvas id="domainChart"></canvas></div>
  </div>
  <div class="card">
    <div class="card-head"><div class="card-title">Recent changes (last 30 days)</div><span class="card-link" id="recent-count"></span></div>
    <div id="recent-feed" style="max-height:280px; overflow-y:auto;"></div>
  </div>
</div>

<div class="card">
  <div class="card-head"><div class="card-title">All regulations — filter &amp; search</div></div>
  <div class="controls">
    <input id="search" type="text" placeholder="Search by name, ID, jurisdiction…" />
    <select id="filter-tab"><option value="">All tabs</option></select>
    <select id="filter-juris"><option value="">All jurisdictions</option></select>
    <select id="filter-risk"><option value="">All risk levels</option><option value="RED">RED only</option><option value="YELLOW">YELLOW only</option><option value="GREEN">GREEN only</option></select>
    <select id="filter-status"><option value="">All statuses</option></select>
    <span class="count" id="reg-count"></span>
  </div>
  <div style="max-height:500px; overflow-y:auto;">
    <table class="reg-table" id="reg-table"><thead><tr><th>ID</th><th>Name</th><th>Jurisdiction</th><th>Domain</th><th>Status</th><th>Effective</th><th>Risk</th></tr></thead><tbody></tbody></table>
  </div>
</div>

<div class="footer">HR Compliance Framework · Generated __TODAY__ · Source: HR_Compliance_Framework_COMPLETE.xlsx · Daily scans 3:06 AM Pacific (Mon–Fri) · Primary federal &amp; state sources only</div>
</div>

<div class="reg-tooltip" id="reg-tooltip"></div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
const DATA = __DATA_JSON__;

// ============ State Panel (used by map and federal button) ============
function showStatePanel(stateName) {
  const panel = document.getElementById('state-panel');
  const title = document.getElementById('state-panel-title');
  const body = document.getElementById('state-panel-body');
  const sd = DATA.state_data[stateName];
  if (!sd || sd.count === 0) {
    title.textContent = stateName + ' — no regulations tracked';
    body.innerHTML = '<div style="color:var(--text2);font-size:13px;">No regulations are currently tracked for this jurisdiction.</div>';
  } else {
    title.textContent = stateName + ' — ' + sd.count + ' regulation' + (sd.count === 1 ? '' : 's');
    body.innerHTML = '';
    const domainOrder = Object.keys(sd.by_domain).sort((a, b) => sd.by_domain[b].length - sd.by_domain[a].length);
    domainOrder.forEach(domain => {
      const regs = sd.by_domain[domain];
      const group = document.createElement('div'); group.className = 'domain-group';
      const head = document.createElement('div'); head.className = 'domain-group-head';
      head.textContent = domain + ' (' + regs.length + ')';
      group.appendChild(head);
      regs.forEach(r => {
        const row = document.createElement('div'); row.className = 'domain-reg';
        if (r.law_id) row.setAttribute('data-summary', r.law_id);
        const left = document.createElement('div'); left.style.flex = '1';
        const nameHtml = r.url ? `<a href="${r.url}" target="_blank" rel="noopener">${r.name}</a>` : r.name;
        left.innerHTML = `<div class="domain-reg-name">${nameHtml}<span class="domain-reg-id">${r.law_id}</span></div>
                          <div class="domain-reg-meta">${r.status || ''}${r.effective ? ' · eff. ' + r.effective : ''}</div>`;
        const owner = document.createElement('span');
        const isOwned = r.owner && r.owner !== '—';
        owner.className = 'coe' + (isOwned ? '' : ' unowned');
        owner.textContent = isOwned ? r.owner : 'No CoE owner';
        row.appendChild(left);
        row.appendChild(owner);
        if (r.risk && r.risk !== 'UNKNOWN') {
          const pill = document.createElement('span');
          pill.className = 'pill ' + r.risk;
          pill.style.marginLeft = '6px';
          pill.textContent = r.risk;
          row.appendChild(pill);
        }
        group.appendChild(row);
      });
      body.appendChild(group);
    });
  }
  panel.classList.add('open');
  panel.scrollIntoView({ behavior: 'smooth', block: 'nearest' });
}
document.getElementById('state-panel-close').addEventListener('click', () => {
  document.getElementById('state-panel').classList.remove('open');
});

// ============ Federal button ============
const fedCount = (DATA.state_data['Federal'] || {count: 0}).count;
document.getElementById('federal-count').textContent = fedCount;
document.getElementById('federal-btn').addEventListener('click', () => showStatePanel('Federal'));

// ============ US Pill Map (fully offline, no CDN) ============
// US country outline (Alaska + lower 48 + Hawaii) as a static SVG path. Coords roughly match
// AlbersUSA in a 980x620 viewBox. Then transparent number pills are positioned at each state's
// centroid. Tiny NE states use leader-line callouts to pills along the right margin.

const US_OUTLINE_PATH = "M 234.0,136.0 L 234.7,133.3 L 251.2,138.0 L 252.2,136.3 L 253.6,139.4 L 248.6,145.7 L 254.4,139.1 L 254.2,146.6 L 251.3,146.8 L 253.0,148.8 L 256.2,145.8 L 255.3,142.1 L 257.1,138.4 L 252.7,132.4 L 255.3,130.8 L 252.2,126.0 L 507.8,126.0 L 507.8,121.4 L 509.7,121.6 L 512.6,129.1 L 520.4,131.8 L 527.9,130.5 L 541.0,137.5 L 547.3,135.1 L 560.3,137.8 L 547.5,143.3 L 536.5,152.3 L 537.6,153.8 L 547.5,150.5 L 546.7,154.8 L 551.5,155.3 L 575.9,144.3 L 569.5,152.9 L 572.7,150.4 L 571.4,152.0 L 575.8,151.3 L 580.1,156.0 L 585.0,156.7 L 602.3,152.8 L 601.6,156.2 L 609.9,155.6 L 608.8,159.9 L 616.0,162.2 L 605.0,161.4 L 604.2,164.0 L 597.1,160.9 L 590.0,162.7 L 586.9,166.7 L 587.6,163.3 L 583.7,165.9 L 582.8,164.1 L 573.7,179.2 L 583.5,170.5 L 574.9,195.2 L 575.6,206.4 L 580.3,214.5 L 586.0,212.2 L 590.2,204.7 L 590.5,197.8 L 587.6,190.1 L 590.2,177.7 L 596.8,171.5 L 597.0,177.0 L 598.4,170.8 L 602.6,169.3 L 601.7,164.9 L 616.8,170.8 L 617.3,181.9 L 611.5,189.0 L 614.1,190.9 L 621.2,185.2 L 623.7,187.8 L 624.8,202.7 L 622.3,202.2 L 616.2,213.5 L 625.2,217.4 L 656.0,203.4 L 658.8,200.6 L 656.8,194.9 L 676.4,194.8 L 683.0,191.6 L 684.0,187.2 L 682.3,183.6 L 694.7,174.2 L 726.8,173.8 L 728.8,170.4 L 732.8,171.2 L 738.3,163.3 L 740.8,153.7 L 747.9,144.5 L 750.9,147.9 L 757.1,145.7 L 761.2,149.2 L 761.1,165.9 L 764.4,166.8 L 764.0,170.6 L 768.7,176.3 L 757.5,183.1 L 754.8,180.8 L 753.7,185.9 L 752.0,180.1 L 750.0,180.9 L 749.3,186.3 L 739.2,189.9 L 735.6,194.1 L 733.2,199.8 L 735.2,202.1 L 731.4,206.6 L 736.9,213.1 L 740.5,212.6 L 738.5,209.1 L 740.4,209.8 L 740.8,215.5 L 732.0,217.1 L 734.7,215.5 L 734.1,213.2 L 728.3,216.6 L 729.7,213.9 L 727.9,212.6 L 727.0,217.5 L 713.8,218.5 L 707.0,222.0 L 705.7,224.5 L 723.6,221.2 L 711.0,226.4 L 703.2,225.9 L 701.2,228.1 L 703.9,228.6 L 702.8,236.9 L 695.7,246.7 L 695.5,244.1 L 689.5,240.5 L 689.5,237.7 L 693.9,253.0 L 685.7,268.9 L 688.3,258.4 L 686.2,259.0 L 686.5,253.7 L 682.8,254.1 L 682.6,250.7 L 684.9,251.2 L 682.0,249.8 L 683.3,246.9 L 681.7,247.8 L 685.5,239.3 L 679.9,242.8 L 682.2,257.6 L 673.6,253.4 L 673.6,250.4 L 673.5,254.0 L 675.5,253.2 L 683.0,259.3 L 681.8,262.7 L 676.8,258.3 L 682.4,263.3 L 682.6,266.3 L 680.6,267.0 L 682.4,270.0 L 676.4,267.2 L 680.7,271.2 L 685.2,270.9 L 689.5,284.6 L 685.2,275.7 L 687.0,281.1 L 683.5,278.4 L 684.6,280.1 L 679.9,281.9 L 678.8,278.7 L 679.0,282.8 L 684.6,282.1 L 684.8,286.0 L 687.0,282.5 L 687.7,286.5 L 683.9,290.1 L 680.7,289.5 L 681.0,287.4 L 679.8,289.4 L 675.7,287.8 L 680.9,290.9 L 676.1,293.9 L 682.5,293.9 L 679.0,297.5 L 680.4,298.2 L 684.8,293.3 L 680.2,298.9 L 675.6,298.1 L 670.8,301.0 L 667.0,307.8 L 659.2,309.1 L 654.1,317.9 L 641.2,326.3 L 637.0,333.4 L 634.8,348.0 L 647.9,392.4 L 644.8,410.1 L 646.5,408.2 L 642.1,415.2 L 644.1,411.7 L 637.6,412.0 L 636.2,405.7 L 632.5,403.8 L 628.6,396.2 L 628.9,390.5 L 627.1,393.4 L 622.8,383.5 L 625.8,379.2 L 623.3,377.7 L 622.9,382.6 L 621.8,379.6 L 623.3,367.1 L 614.5,355.4 L 608.6,352.8 L 601.4,359.0 L 589.8,349.6 L 574.1,351.3 L 576.3,350.5 L 574.0,345.8 L 572.8,350.2 L 561.9,349.5 L 556.9,353.9 L 562.8,354.3 L 558.5,358.5 L 564.8,363.8 L 560.9,366.8 L 562.2,363.8 L 557.0,362.2 L 553.5,365.0 L 550.2,362.3 L 547.5,365.3 L 538.6,356.1 L 534.7,359.5 L 523.4,356.8 L 511.8,361.6 L 513.9,359.7 L 509.8,357.6 L 510.3,362.3 L 507.7,364.6 L 511.8,362.0 L 493.3,376.0 L 488.8,382.8 L 487.2,389.3 L 489.3,401.0 L 487.4,385.8 L 490.0,379.0 L 496.1,373.1 L 489.3,376.4 L 489.0,380.0 L 486.0,379.8 L 488.4,381.6 L 484.8,386.8 L 486.8,386.8 L 486.6,394.7 L 489.3,401.3 L 487.3,403.9 L 471.5,397.2 L 467.5,383.2 L 449.9,356.9 L 438.2,357.1 L 436.2,363.5 L 432.6,366.2 L 421.2,358.3 L 417.4,346.7 L 403.2,332.8 L 387.0,332.6 L 387.0,338.0 L 360.4,338.0 L 325.8,324.1 L 326.7,321.4 L 304.4,323.6 L 300.9,314.0 L 292.5,309.1 L 291.3,305.5 L 272.0,299.4 L 271.9,292.5 L 260.4,278.5 L 260.5,270.8 L 255.5,267.7 L 254.5,260.6 L 258.8,264.5 L 255.3,258.4 L 256.5,256.7 L 254.7,256.7 L 254.3,260.2 L 249.9,258.0 L 250.3,255.0 L 243.1,246.6 L 242.1,236.0 L 237.4,230.9 L 240.2,216.7 L 235.6,199.9 L 239.4,187.1 L 240.7,159.2 L 244.9,158.8 L 240.0,158.8 L 240.1,154.4 L 241.2,157.5 L 242.3,153.4 L 239.5,151.1 L 242.0,150.6 L 239.1,150.9 L 234.0,136.0 Z M 23.1,561.0 L 26.6,559.3 L 23.1,561.0 Z M 19.6,560.3 L 22.8,558.6 L 19.6,560.3 Z M 17.1,558.8 L 20.0,559.0 L 17.1,558.8 Z M 243.5,530.4 L 245.0,529.0 L 243.5,530.4 Z M 236.6,518.5 L 238.8,518.2 L 236.6,518.5 Z M 233.1,526.6 L 235.6,525.9 L 233.1,526.6 Z M 233.0,522.0 L 236.2,518.8 L 242.0,532.6 L 236.4,528.4 L 237.6,533.4 L 233.6,523.7 L 235.4,523.3 L 233.0,522.0 Z M 230.1,514.1 L 236.5,512.7 L 239.2,516.4 L 233.5,516.7 L 231.4,521.6 L 230.1,514.1 Z M 227.4,500.1 L 231.2,502.3 L 232.7,509.5 L 229.2,512.4 L 227.4,500.1 Z M 223.9,509.4 L 228.0,509.0 L 228.9,520.2 L 223.9,509.4 Z M 223.1,510.5 L 225.3,510.6 L 223.1,510.5 Z M 219.6,504.4 L 223.4,501.1 L 227.4,503.3 L 228.1,508.2 L 222.9,509.0 L 219.6,504.4 Z M 164.4,487.6 L 169.3,483.1 L 164.4,487.6 Z M 164.3,483.5 L 166.1,480.5 L 164.3,483.5 Z M 137.7,503.1 L 141.8,498.1 L 144.8,500.7 L 137.7,503.1 Z M 133.1,516.8 L 135.5,516.6 L 133.1,516.8 Z M 131.1,509.4 L 138.3,503.7 L 143.8,507.1 L 135.0,515.0 L 135.8,511.3 L 133.4,514.1 L 131.1,509.4 Z M 130.9,517.8 L 133.0,516.7 L 130.9,517.8 Z M 118.7,516.8 L 120.3,516.9 L 118.7,516.8 Z M 104.5,531.3 L 106.4,528.3 L 104.5,531.3 Z M 101.5,527.8 L 104.2,528.0 L 101.5,527.8 Z M 100.4,498.4 L 102.4,496.4 L 100.4,498.4 Z M 96.4,529.6 L 98.2,530.1 L 96.4,529.6 Z M 92.6,453.4 L 94.3,453.4 L 92.6,453.4 Z M 91.9,535.6 L 94.2,536.2 L 91.9,535.6 Z M 80.3,538.8 L 81.8,538.7 L 80.3,538.8 Z M 76.0,538.4 L 78.1,538.5 L 76.0,538.4 Z M 69.5,483.8 L 75.8,481.8 L 78.7,485.9 L 75.6,487.9 L 69.5,483.8 Z M 67.5,545.9 L 72.6,542.7 L 71.5,540.2 L 76.1,541.1 L 67.5,545.9 Z M 292.2,536.8 L 294.0,534.7 L 293.5,532.4 L 294.2,532.2 L 299.4,534.8 L 300.4,535.8 L 300.3,536.9 L 301.0,536.8 L 302.7,538.8 L 301.2,540.4 L 296.7,542.3 L 295.4,544.2 L 293.6,543.1 L 293.6,540.4 L 292.2,536.8 Z M 286.7,526.2 L 287.7,525.2 L 288.7,526.4 L 290.7,526.1 L 292.8,528.0 L 289.5,529.3 L 288.7,527.5 L 287.3,527.1 L 286.7,526.2 Z M 286.7,529.7 L 288.1,529.3 L 287.7,529.9 L 286.7,529.7 Z M 283.7,526.4 L 285.1,526.3 L 285.8,527.1 L 284.5,527.9 L 283.7,526.4 Z M 281.6,524.6 L 282.0,523.5 L 286.7,524.1 L 285.3,525.1 L 281.6,524.6 Z M 273.4,520.3 L 274.7,520.2 L 276.1,519.1 L 277.1,521.4 L 278.1,521.4 L 278.7,522.8 L 274.8,522.8 L 273.4,520.3 Z M 260.7,516.2 L 262.4,514.5 L 264.0,514.4 L 264.9,515.2 L 264.6,516.9 L 263.6,517.7 L 260.7,516.2 Z M 256.9,517.9 L 258.3,516.5 L 258.3,517.4 L 257.2,518.4 L 256.9,517.9 Z";

const STATES_GEO = {
  'Alabama': [585.3, 320.3, 'AL'],
  'Alaska': [142.6, 447.9, 'AK'],
  'Arizona': [355.0, 304.0, 'AZ'],
  'Arkansas': [533.0, 297.0, 'AR'],
  'California': [280.7, 272.6, 'CA'],
  'Colorado': [411.6, 246.0, 'CO'],
  'Connecticut': [715.5, 214.6, 'CT'],
  'Delaware': [689.8, 246.1, 'DE'],
  'District of Columbia': [675.8, 247.1, 'DC'],
  'Florida': [634.6, 382.1, 'FL'],
  'Georgia': [614.4, 322.2, 'GA'],
  'Hawaii': [280.0, 530.0, 'HI'],
  'Idaho': [326.5, 183.1, 'ID'],
  'Illinois': [563.0, 233.5, 'IL'],
  'Indiana': [590.2, 235.3, 'IN'],
  'Iowa': [525.8, 209.2, 'IA'],
  'Kansas': [478.0, 252.1, 'KS'],
  'Kentucky': [599.4, 259.9, 'KY'],
  'Louisiana': [537.4, 341.2, 'LA'],
  'Maine': [746.5, 169.6, 'ME'],
  'Maryland': [677.8, 245.3, 'MD'],
  'Massachusetts': [723.8, 207.2, 'MA'],
  'Michigan': [606.1, 181.9, 'MI'],
  'Minnesota': [515.6, 158.6, 'MN'],
  'Mississippi': [558.7, 321.1, 'MS'],
  'Missouri': [532.8, 253.7, 'MO'],
  'Montana': [366.2, 151.0, 'MT'],
  'Nebraska': [464.9, 215.5, 'NE'],
  'Nevada': [308.8, 254.3, 'NV'],
  'New Hampshire': [726.2, 192.6, 'NH'],
  'New Jersey': [698.9, 230.4, 'NJ'],
  'New Mexico': [406.4, 301.1, 'NM'],
  'New York': [689.8, 198.6, 'NY'],
  'North Carolina': [649.9, 287.4, 'NC'],
  'North Dakota': [460.0, 144.6, 'ND'],
  'Ohio': [622.3, 230.5, 'OH'],
  'Oklahoma': [486.2, 290.3, 'OK'],
  'Oregon': [272.7, 185.2, 'OR'],
  'Pennsylvania': [674.0, 226.9, 'PA'],
  'Rhode Island': [727.6, 213.8, 'RI'],
  'South Carolina': [637.4, 307.7, 'SC'],
  'South Dakota': [460.8, 182.4, 'SD'],
  'Tennessee': [586.2, 283.7, 'TN'],
  'Texas': [474.9, 341.4, 'TX'],
  'Utah': [355.9, 242.2, 'UT'],
  'Vermont': [715.7, 185.2, 'VT'],
  'Virginia': [665.1, 260.8, 'VA'],
  'Washington': [270.9, 145.2, 'WA'],
  'West Virginia': [639.4, 252.1, 'WV'],
  'Wisconsin': [555.6, 181.3, 'WI'],
  'Wyoming': [395.4, 198.1, 'WY'],
};

// States that need leader-line callouts (small/crowded). Callout target is right margin.
const CALLOUTS = {
  'Vermont':                 { from: [715.7, 185.2], to: [920, 120] },
  'New Hampshire':           { from: [726.2, 192.6], to: [920, 152] },
  'Massachusetts':           { from: [723.8, 207.2], to: [920, 184] },
  'Rhode Island':            { from: [727.6, 213.8], to: [920, 216] },
  'Connecticut':             { from: [715.5, 214.6], to: [920, 248] },
  'New Jersey':              { from: [698.9, 230.4], to: [920, 280] },
  'Delaware':                { from: [689.8, 246.1], to: [920, 312] },
  'District of Columbia':    { from: [675.8, 247.1], to: [920, 344] },
  'Maryland':                { from: [677.8, 245.3], to: [920, 376] },
};

let currentCategoryFilter = '';

function getFilteredStateCounts() {
  const states = DATA.state_data;
  const counts = {};
  Object.keys(states).forEach(name => {
    if (!currentCategoryFilter) {
      counts[name] = states[name].count;
    } else {
      let n = 0;
      Object.values(states[name].by_domain || {}).forEach(regs => {
        regs.forEach(r => { if ((r.category_norm || '') === currentCategoryFilter) n++; });
      });
      counts[name] = n;
    }
  });
  return counts;
}

function getFilteredStatePanel(stateName) {
  // Returns a filtered copy of states[stateName] containing only regs in currentCategoryFilter
  const sd = DATA.state_data[stateName];
  if (!sd) return { count: 0, by_domain: {} };
  if (!currentCategoryFilter) return sd;
  const filtered = { count: 0, by_domain: {} };
  Object.entries(sd.by_domain).forEach(([domain, regs]) => {
    const matched = regs.filter(r => (r.category_norm || '') === currentCategoryFilter);
    if (matched.length) {
      filtered.by_domain[domain] = matched;
      filtered.count += matched.length;
    }
  });
  return filtered;
}

function renderPillMap() {
  const stateCounts = getFilteredStateCounts();
  const nonFedCounts = Object.entries(stateCounts).filter(([n]) => n !== 'Federal').map(([_,v]) => v);
  const maxCount = Math.max(1, ...nonFedCounts);

  function tierFor(c) {
    if (!c) return 0;
    const t = c / maxCount;
    if (t >= 0.66) return 3;
    if (t >= 0.33) return 2;
    return 1;
  }

  const svg = document.getElementById('map-svg');
  const NS = 'http://www.w3.org/2000/svg';
  // Clear all and re-add the US outline as the first child
  while (svg.firstChild) svg.removeChild(svg.firstChild);
  const outline = document.createElementNS(NS, 'path');
  outline.setAttribute('d', US_OUTLINE_PATH);
  outline.setAttribute('class', 'us-outline');
  svg.appendChild(outline);

  // Callout lines
  Object.entries(CALLOUTS).forEach(([name, {from, to}]) => {
    const line = document.createElementNS(NS, 'path');
    const elbow = [to[0] - 25, from[1]];
    const d = 'M ' + from[0] + ',' + from[1] + ' L ' + elbow[0] + ',' + elbow[1] + ' L ' + to[0] + ',' + to[1];
    line.setAttribute('d', d);
    line.setAttribute('class', 'callout-line');
    svg.appendChild(line);
  });

  // Pills
  Object.entries(STATES_GEO).forEach(([name, [x, y, abbr]]) => {
    const callout = CALLOUTS[name];
    const px = callout ? callout.to[0] : x;
    const py = callout ? callout.to[1] : y;
    const count = stateCounts[name] || 0;
    const tier = tierFor(count);
    const g = document.createElementNS(NS, 'g');
    g.setAttribute('class', 'pill-group');
    g.setAttribute('transform', 'translate(' + px + ',' + py + ')');
    const bg = document.createElementNS(NS, 'rect');
    bg.setAttribute('class', 'pill-bg');
    bg.setAttribute('x', -16); bg.setAttribute('y', -14);
    bg.setAttribute('width', 32); bg.setAttribute('height', 28);
    bg.setAttribute('rx', 4);
    g.appendChild(bg);
    const abbrEl = document.createElementNS(NS, 'text');
    abbrEl.setAttribute('class', 'pill-abbr');
    abbrEl.setAttribute('y', -3);
    abbrEl.textContent = abbr;
    g.appendChild(abbrEl);
    const countEl = document.createElementNS(NS, 'text');
    countEl.setAttribute('class', 'pill-count tier-' + tier);
    countEl.setAttribute('y', 11);
    countEl.textContent = count;
    g.appendChild(countEl);
    const title = document.createElementNS(NS, 'title');
    title.textContent = name + ': ' + count + ' regulation' + (count === 1 ? '' : 's');
    g.appendChild(title);
    if (count > 0) {
      g.addEventListener('click', () => showStatePanel(name));
    } else {
      g.style.cursor = 'default';
    }
    svg.appendChild(g);
  });

  // Federal button count update
  const fedCount = stateCounts['Federal'] || 0;
  document.getElementById('federal-count').textContent = fedCount;
}

// Override showStatePanel to use filtered data
const originalShowStatePanel = showStatePanel;
showStatePanel = function(stateName) {
  const sd = getFilteredStatePanel(stateName);
  const panel = document.getElementById('state-panel');
  const title = document.getElementById('state-panel-title');
  const body = document.getElementById('state-panel-body');
  const filterSuffix = currentCategoryFilter ? ' · category: ' + currentCategoryFilter : '';
  if (sd.count === 0) {
    title.textContent = stateName + ' — no regulations' + filterSuffix;
    body.innerHTML = '<div style="color:var(--text2);font-size:13px;">No regulations match this filter for ' + stateName + '.</div>';
  } else {
    title.textContent = stateName + ' — ' + sd.count + ' regulation' + (sd.count === 1 ? '' : 's') + filterSuffix;
    body.innerHTML = '';
    const domainOrder = Object.keys(sd.by_domain).sort((a, b) => sd.by_domain[b].length - sd.by_domain[a].length);
    domainOrder.forEach(domain => {
      const regs = sd.by_domain[domain];
      const group = document.createElement('div'); group.className = 'domain-group';
      const head = document.createElement('div'); head.className = 'domain-group-head';
      head.textContent = domain + ' (' + regs.length + ')';
      group.appendChild(head);
      regs.forEach(r => {
        const row = document.createElement('div'); row.className = 'domain-reg';
        if (r.law_id) row.setAttribute('data-summary', r.law_id);
        const left = document.createElement('div'); left.style.flex = '1';
        const nameHtml = r.url ? '<a href="' + r.url + '" target="_blank" rel="noopener">' + r.name + '</a>' : r.name;
        const catLabel = r.category ? ' · ' + r.category : '';
        left.innerHTML = '<div class="domain-reg-name">' + nameHtml + '<span class="domain-reg-id">' + r.law_id + '</span></div>' +
                         '<div class="domain-reg-meta">' + (r.status || '') + (r.effective ? ' · eff. ' + r.effective : '') + catLabel + '</div>';
        const owner = document.createElement('span');
        const isOwned = r.owner && r.owner !== '—';
        owner.className = 'coe' + (isOwned ? '' : ' unowned');
        owner.textContent = isOwned ? r.owner : 'No CoE owner';
        row.appendChild(left);
        row.appendChild(owner);
        if (r.risk && r.risk !== 'UNKNOWN') {
          const pill = document.createElement('span');
          pill.className = 'pill ' + r.risk;
          pill.style.marginLeft = '6px';
          pill.textContent = r.risk;
          row.appendChild(pill);
        }
        group.appendChild(row);
      });
      body.appendChild(group);
    });
  }
  panel.classList.add('open');
  panel.scrollIntoView({ behavior: 'smooth', block: 'nearest' });
};

// Build category dropdown options
(function buildCategoryFilter(){
  const sel = document.getElementById('category-filter');
  const cats = DATA.categories || [];
  cats.forEach(c => {
    const opt = document.createElement('option');
    opt.value = c.name;
    opt.textContent = c.name + ' (' + c.count + ')';
    sel.appendChild(opt);
  });
  sel.addEventListener('change', () => {
    currentCategoryFilter = sel.value;
    renderPillMap();
    // If state panel is open, refresh its contents
    const panel = document.getElementById('state-panel');
    if (panel.classList.contains('open')) {
      const titleText = document.getElementById('state-panel-title').textContent;
      const stateName = titleText.split(' — ')[0];
      if (stateName && DATA.state_data[stateName]) showStatePanel(stateName);
    }
  });
})();

renderPillMap();

// ============ Today's changes ============
(function renderToday(){
  const card = document.getElementById('today-card');
  const titleEl = document.getElementById('today-title');
  const badgesEl = document.getElementById('today-badges');
  const summaryEl = document.getElementById('today-summary');
  const bodyEl = document.getElementById('today-body');
  const u = DATA.latest_updates;
  if (!u || !u.tab_date) {
    titleEl.textContent = "Today's changes";
    card.classList.add('empty');
    bodyEl.innerHTML = '<div class="today-empty">No daily update tabs found yet. First scheduled scan runs 3:06 AM Pacific.</div>';
    return;
  }
  const [y,mo,da] = u.tab_date.split('-').map(Number);
  const tabDate = new Date(Date.UTC(y, mo-1, da));
  const today = new Date(); today.setUTCHours(0,0,0,0);
  const ageDays = Math.round((today - tabDate) / 86400000);
  const ageStr = ageDays === 0 ? 'today' : ageDays === 1 ? 'yesterday' : ageDays + ' days ago';
  titleEl.innerHTML = `Today's changes <span style="color:var(--text3);font-weight:400;font-size:12px;">· scan from ${u.tab_date} (${ageStr})</span>`;
  if (ageDays > 2) card.classList.add('stale');
  const material = u.rows.filter(r => {
    const t = (r.type || '').toUpperCase();
    return t && !t.includes('OBSERVATION') && !t.includes('AUDIT');
  });
  const obsCount = u.rows.length - material.length;
  if (material.length === 0) {
    bodyEl.innerHTML = `<div class="today-empty">No regulatory changes in the most recent scan.${obsCount ? ' (' + obsCount + ' observation/audit ' + (obsCount===1?'entry':'entries') + ' logged.)' : ''}</div>`;
    return;
  }

  // ===== Type counts → badges =====
  const typeCounts = {};
  material.forEach(r => {
    const t = (r.type || 'UNKNOWN').toUpperCase().replace(/[^A-Z\-]/g, '');
    typeCounts[t] = (typeCounts[t] || 0) + 1;
  });
  badgesEl.innerHTML = Object.entries(typeCounts).map(([t,n]) => `<span class="tag ${t}">${n} ${t}</span>`).join('');

  // ===== Executive summary =====
  const fedCount = material.filter(r => (r.jurisdiction || '').toLowerCase().includes('federal')).length;
  const stateCount = material.length - fedCount;
  function displayType(t) { return t === 'CATCHUP' ? 'CATCH-UP' : t; }
  const typeStr = Object.entries(typeCounts).map(([t,n]) => `${n} ${displayType(t)}`).join(', ');
  const jurisStr = (fedCount && stateCount) ? `${fedCount} federal, ${stateCount} state` : (fedCount ? `${fedCount} federal` : `${stateCount} state`);

  // Rank highlights: HIGH confidence first, then NEW > CHANGE > SPLIT > CATCH-UP > UNKNOWN
  const TYPE_PRIORITY = { 'NEW': 5, 'CHANGE': 4, 'SPLIT': 3, 'CATCHUP': 2, 'CATCH-UP': 2 };
  const CONF_PRIORITY = { 'HIGH': 3, 'MED': 2, 'LOW': 1 };
  function score(r) {
    const t = (r.type || '').toUpperCase().replace(/[^A-Z\-]/g, '');
    const c = (r.confidence || '').toUpperCase().replace(/[^A-Z]/g, '').slice(0,4);
    return (CONF_PRIORITY[c] || 0) * 10 + (TYPE_PRIORITY[t] || 1);
  }
  const ranked = material.slice().sort((a, b) => score(b) - score(a));
  const highlights = ranked.slice(0, 3);

  // Build one-line framing per highlight
  function frame(r) {
    // Strip wrapping noise; first sentence of `after` or `notes`, capped
    let txt = (r.after || r.notes || '').trim();
    // Try to grab first meaningful sentence
    const firstSentence = txt.split(/(?<=[.!?])\s+/)[0] || txt;
    const trimmed = firstSentence.length > 160 ? firstSentence.slice(0, 157) + '…' : firstSentence;
    return trimmed;
  }

  let summaryHtml = '<div class="today-summary">';
  summaryHtml += `<div class="today-summary-line"><strong>${material.length} regulatory change${material.length===1?'':'s'}</strong> in this scan — ${jurisStr} (${typeStr}).</div>`;
  summaryHtml += '<div class="today-highlights">';
  summaryHtml += '<div class="today-highlights-label">Top highlights</div>';
  highlights.forEach(r => {
    const t = (r.type || 'UNKNOWN').toUpperCase().replace(/[^A-Z\-]/g, '');
    const jur = r.jurisdiction || 'Unknown';
    const name = r.name || '(unnamed)';
    const framing = frame(r);
    const url = r.url || '';
    const nameHtml = url ? `<a href="${url}" target="_blank" rel="noopener">${name}</a>` : name;
    summaryHtml += `<div class="today-highlight">
      <span class="today-highlight-tag tag ${t}">${t}</span>
      <span class="today-highlight-text"><strong>${jur}</strong> · ${nameHtml}${framing ? ` <span class="frame">— ${framing}</span>` : ''}</span>
    </div>`;
  });
  summaryHtml += '</div></div>';
  summaryEl.innerHTML = summaryHtml;

  // ===== Detail list (existing behavior) =====
  const list = document.createElement('div'); list.className = 'today-list';
  material.forEach(r => {
    const t = (r.type || 'UNKNOWN').toUpperCase().replace(/[^A-Z\-]/g, '');
    const conf = (r.confidence || '').toUpperCase().replace(/[^A-Z]/g, '').slice(0,4) || 'UNKNOWN';
    const item = document.createElement('div'); item.className = 'today-item';
    const name = r.name || '(unnamed)';
    const lawId = r.law_id ? r.law_id + ' · ' : '';
    const jur = r.jurisdiction ? r.jurisdiction + ' · ' : '';
    const tab = r.affected_tab || '';
    const fieldTxt = r.field ? ' · ' + r.field : '';
    const url = r.url || '';
    if (r.law_id) item.setAttribute('data-summary', r.law_id);
    item.innerHTML = `
      <div class="today-item-head">
        <span class="tag ${t}">${t}</span>
        ${(conf === 'HIGH' || conf === 'MED' || conf === 'LOW') ? `<span class="tag ${conf}">${conf}</span>` : ''}
        <span class="today-item-name">${url ? `<a href="${url}" target="_blank" rel="noopener">${name}</a>` : name}</span>
      </div>
      <div class="today-item-meta">${lawId}${jur}${tab}${fieldTxt}</div>
      <div class="today-item-detail">
        ${r.before ? `<div class="diff-row"><div class="diff-label">Before</div><div class="diff-text">${r.before}</div></div>` : ''}
        ${r.after ? `<div class="diff-row"><div class="diff-label">After</div><div class="diff-text">${r.after}</div></div>` : ''}
        ${r.notes ? `<div class="diff-row"><div class="diff-label">Notes</div><div class="diff-text">${r.notes}</div></div>` : ''}
      </div>`;
    item.addEventListener('click', e => { if (e.target.tagName !== 'A') item.classList.toggle('open'); });
    list.appendChild(item);
  });
  bodyEl.appendChild(list);
  if (obsCount) {
    const note = document.createElement('div');
    note.style.cssText = 'font-size:11px;color:var(--text3);margin-top:10px;padding-top:8px;border-top:0.5px solid var(--border);';
    note.textContent = '+ ' + obsCount + ' observation/audit ' + (obsCount===1?'entry':'entries') + ' in this scan.';
    bodyEl.appendChild(note);
  }
})();

// ============ Upcoming + charts + recent + reg table (unchanged) ============
const fmt = (iso) => { const [y,m,d] = iso.split('-'); return m+'/'+d+'/'+y.slice(2); };
const upTbody = document.querySelector('#upcoming-table tbody');
DATA.metrics.effective_next_12mo.forEach(e => {
  const tr = document.createElement('tr');
  tr.setAttribute('data-summary', e.law_id);
  tr.innerHTML = `<td style="color:var(--text2);white-space:nowrap;">${fmt(e.effective)}</td>
                  <td>${e.url ? `<a href="${e.url}" target="_blank" rel="noopener">${e.name}</a>` : e.name}<br><span style="color:var(--text3);font-size:10px;">${e.law_id}</span></td>
                  <td>${e.jurisdiction}</td>
                  <td style="text-align:right;"><span class="pill ${e.risk}">${e.risk}</span></td>`;
  upTbody.appendChild(tr);
});
document.getElementById('upcoming-count').textContent = DATA.metrics.effective_next_12mo.length + ' items';
document.getElementById('recent-count').textContent = DATA.metrics.recent_changes_count + ' items';

const jurisEntries = Object.entries(DATA.metrics.by_jurisdiction);
new Chart(document.getElementById('jurisChart'), {
  type: 'bar',
  data: { labels: jurisEntries.map(e=>e[0]), datasets: [{ label: 'Regs', data: jurisEntries.map(e=>e[1]), backgroundColor: '#534AB7', borderRadius: 4 }] },
  options: { indexAxis:'y', responsive:true, maintainAspectRatio:false, plugins:{ legend:{display:false} }, scales:{ x:{beginAtZero:true,ticks:{font:{size:11}}}, y:{ticks:{font:{size:11}}} } }
});
const domainEntries = Object.entries(DATA.metrics.by_domain);
const colors = ['#534AB7','#1D9E75','#D85A30','#378ADD','#BA7517','#888780','#D4537E','#639922'];
new Chart(document.getElementById('domainChart'), {
  type: 'doughnut',
  data: { labels: domainEntries.map(e=>e[0]), datasets: [{ data: domainEntries.map(e=>e[1]), backgroundColor: domainEntries.map((_,i)=>colors[i%colors.length]), borderWidth: 0 }] },
  options: { responsive:true, maintainAspectRatio:false, cutout:'65%', plugins:{ legend:{display:false} } }
});
const legendDiv = document.getElementById('domain-legend');
domainEntries.forEach(([name, count], i) => {
  const span = document.createElement('span');
  span.innerHTML = `<span class="legend-sw" style="background:${colors[i%colors.length]}"></span>${name} ${count}`;
  legendDiv.appendChild(span);
});
const feedDiv = document.getElementById('recent-feed');
if (DATA.metrics.recent_changes.length === 0) {
  feedDiv.innerHTML = '<div style="color:var(--text3);font-size:12px;padding:12px 0;">No changes in last 30 days.</div>';
} else {
  DATA.metrics.recent_changes.forEach(c => {
    const div = document.createElement('div');
    div.className = 'feed-item';
    div.setAttribute('data-summary', c.law_id);
    div.innerHTML = `<span class="feed-tag">${(c.status || 'UPDATE').slice(0,30)}</span>
                     <div><div>${c.name}</div><div class="feed-meta">${c.law_id} · ${c.jurisdiction} · ${fmt(c.last_review)}</div></div>`;
    feedDiv.appendChild(div);
  });
}

const regs = DATA.rows;
const tabs = [...new Set(regs.map(r => r['__tab__']))].sort();
const jurises = [...new Set(regs.map(r => r['Jurisdiction']).filter(Boolean))].sort();
const statuses = [...new Set(regs.map(r => (r['Status']||'').trim()).filter(Boolean))].sort();
const tabSel = document.getElementById('filter-tab');
const jSel = document.getElementById('filter-juris');
const sSel = document.getElementById('filter-status');
tabs.forEach(t => { const o = document.createElement('option'); o.value = t; o.textContent = t; tabSel.appendChild(o); });
jurises.forEach(j => { const o = document.createElement('option'); o.value = j; o.textContent = j; jSel.appendChild(o); });
statuses.forEach(s => { const o = document.createElement('option'); o.value = s; o.textContent = s; sSel.appendChild(o); });
function normRisk(s) { s = (s||'').toUpperCase(); if (s.includes('RED')) return 'RED'; if (s.includes('YELLOW')) return 'YELLOW'; if (s.includes('GREEN')) return 'GREEN'; return 'UNKNOWN'; }
function renderTable() {
  const q = document.getElementById('search').value.toLowerCase();
  const tab = tabSel.value;
  const juris = jSel.value;
  const risk = document.getElementById('filter-risk').value;
  const status = sSel.value;
  const tbody = document.querySelector('#reg-table tbody');
  tbody.innerHTML = '';
  let shown = 0;
  regs.forEach(r => {
    if (tab && r['__tab__'] !== tab) return;
    if (juris && r['Jurisdiction'] !== juris) return;
    const nr = normRisk(r['Risk Level']);
    if (risk && nr !== risk) return;
    if (status && (r['Status']||'').trim() !== status) return;
    if (q) {
      const hay = (r['Law / Regulation Name']+' '+r['Law ID']+' '+r['Jurisdiction']+' '+r['Status']+' '+r['Core Domain']).toLowerCase();
      if (!hay.includes(q)) return;
    }
    const tr = document.createElement('tr');
    tr.setAttribute('data-summary', r['Law ID']);
    const url = r['Gov Source URL'];
    tr.innerHTML = `<td style="color:var(--text3);white-space:nowrap;">${r['Law ID']}</td>
                    <td>${url ? `<a href="${url}" target="_blank" rel="noopener">${r['Law / Regulation Name']}</a>` : r['Law / Regulation Name']}</td>
                    <td>${r['Jurisdiction']}</td>
                    <td style="color:var(--text2);">${r['Core Domain']||''}</td>
                    <td style="color:var(--text2);">${r['Status']||''}</td>
                    <td style="color:var(--text2);font-size:10px;">${r['Effective Date']||''}</td>
                    <td><span class="pill ${nr}">${nr}</span></td>`;
    tbody.appendChild(tr);
    shown++;
  });
  document.getElementById('reg-count').textContent = shown + ' of ' + regs.length + ' regs';
}
['search','filter-tab','filter-juris','filter-risk','filter-status'].forEach(id => {
  document.getElementById(id).addEventListener('input', renderTable);
  document.getElementById(id).addEventListener('change', renderTable);
});
renderTable();

// ============ Reg tooltip (shows Summary / Key Requirements on hover) ============
(function setupRegTooltip(){
  const tip = document.getElementById('reg-tooltip');
  // Build a fast lookup: lawId -> { name, summary, jurisdiction, status }
  const summaryByLawId = {};
  DATA.rows.forEach(r => {
    const lid = r['Law ID'];
    if (lid && !summaryByLawId[lid]) {
      summaryByLawId[lid] = {
        name: r['Law / Regulation Name'] || '',
        summary: r['Summary / Key Requirements'] || '',
        jurisdiction: r['Jurisdiction'] || '',
        status: r['Status'] || '',
        effective: r['Effective Date'] || '',
      };
    }
  });

  function showTip(el, evt) {
    const lid = el.getAttribute('data-summary');
    if (!lid) return;
    const info = summaryByLawId[lid];
    if (!info) {
      tip.innerHTML = '<div class="rt-empty">No summary available for ' + lid + '</div>';
    } else {
      const summary = info.summary && info.summary.trim() ? info.summary : '<span class="rt-empty">No summary recorded for this regulation.</span>';
      tip.innerHTML = '<div class="rt-title">' + info.name + '</div>' +
                      '<div class="rt-meta">' + lid + ' &middot; ' + (info.jurisdiction || '—') +
                      (info.status ? ' &middot; ' + info.status : '') +
                      (info.effective ? ' &middot; eff. ' + info.effective : '') +
                      '</div>' +
                      '<div class="rt-body">' + summary + '</div>';
    }
    tip.classList.add('visible');
    moveTip(evt);
  }
  function moveTip(evt) {
    // Position near cursor, but flip if it would overflow viewport
    const rect = tip.getBoundingClientRect();
    let x = evt.clientX + 16;
    let y = evt.clientY + 16;
    if (x + rect.width > window.innerWidth - 16) x = evt.clientX - rect.width - 16;
    if (y + rect.height > window.innerHeight - 16) y = evt.clientY - rect.height - 16;
    if (x < 8) x = 8;
    if (y < 8) y = 8;
    tip.style.left = x + 'px';
    tip.style.top = y + 'px';
  }
  function hideTip() {
    tip.classList.remove('visible');
  }
  // Delegate via mouseover/mouseout for performance (catches dynamically-added elements)
  document.addEventListener('mouseover', e => {
    const target = e.target.closest('[data-summary]');
    if (target) showTip(target, e);
  });
  document.addEventListener('mousemove', e => {
    if (tip.classList.contains('visible')) moveTip(e);
  });
  document.addEventListener('mouseout', e => {
    const target = e.target.closest('[data-summary]');
    if (target) hideTip();
  });
})();

// ============ KPI Drill-downs ============
(function setupKpiDrill(){
  const panel = document.getElementById('kpi-drill');
  const titleEl = document.getElementById('kpi-drill-title');
  const bodyEl = document.getElementById('kpi-drill-body');
  const closeEl = document.getElementById('kpi-drill-close');
  let activeKey = null;

  function normRiskLocal(s) { s = (s||'').toUpperCase(); if (s.includes('RED')) return 'RED'; if (s.includes('YELLOW')) return 'YELLOW'; if (s.includes('GREEN')) return 'GREEN'; return 'UNKNOWN'; }
  function pill(risk) { return '<span class="pill ' + risk + '">' + risk + '</span>'; }

  function renderRowsTable(rows, opts) {
    if (!rows.length) {
      return '<div class="kpi-drill-empty">No regulations match this filter.</div>';
    }
    const showEffCol = opts && opts.showEff !== false;
    const showLastReview = opts && opts.lastReview;
    let html = '<table><thead><tr>';
    html += '<th>ID</th><th>Regulation</th><th>Jurisdiction</th><th>Domain</th><th>Status</th>';
    if (showLastReview) html += '<th>Last review</th>';
    else if (showEffCol) html += '<th>Effective</th>';
    html += '<th style="text-align:right;">Risk</th></tr></thead><tbody>';
    rows.forEach(r => {
      const url = r['Gov Source URL'] || r.url || '';
      const name = r['Law / Regulation Name'] || r.name || '(unnamed)';
      const lawId = r['Law ID'] || r.law_id || '';
      const jur = r['Jurisdiction'] || r.jurisdiction || '';
      const dom = r['Core Domain'] || r.domain || '';
      const status = r['Status'] || r.status || '';
      const eff = r['Effective Date'] || r.effective_raw || r.effective || '';
      const risk = normRiskLocal(r['Risk Level'] || r.risk);
      html += '<tr data-summary="' + (lawId || '') + '">';
      html += '<td class="row-id">' + lawId + '</td>';
      html += '<td>' + (url ? '<a href="' + url + '" target="_blank" rel="noopener">' + name + '</a>' : name) + '</td>';
      html += '<td>' + jur + '</td>';
      html += '<td style="color:var(--text2);">' + dom + '</td>';
      html += '<td style="color:var(--text2);">' + status + '</td>';
      if (showLastReview) {
        html += '<td class="row-eff">' + (r.last_review || '') + '</td>';
      } else if (showEffCol) {
        html += '<td class="row-eff">' + eff + '</td>';
      }
      html += '<td style="text-align:right;">' + pill(risk) + '</td>';
      html += '</tr>';
    });
    html += '</tbody></table>';
    return html;
  }

  function openDrill(key, srcEl) {
    activeKey = key;
    // Mark active card
    document.querySelectorAll('.kpi.clickable').forEach(el => el.classList.remove('active'));
    if (srcEl) srcEl.classList.add('active');

    let title, rows, opts = {};
    if (key === 'all') {
      title = 'All regulations (' + DATA.rows.length + ')';
      // Sort: RED first, then by jurisdiction
      rows = DATA.rows.slice().sort((a, b) => {
        const ra = normRiskLocal(a['Risk Level']), rb = normRiskLocal(b['Risk Level']);
        const ord = { RED:0, YELLOW:1, GREEN:2, UNKNOWN:3 };
        if (ord[ra] !== ord[rb]) return ord[ra] - ord[rb];
        return (a['Jurisdiction'] || '').localeCompare(b['Jurisdiction'] || '');
      });
    } else if (key === 'red') {
      rows = DATA.rows.filter(r => normRiskLocal(r['Risk Level']) === 'RED')
                       .sort((a, b) => (a['Jurisdiction'] || '').localeCompare(b['Jurisdiction'] || ''));
      title = 'RED risk regulations (' + rows.length + ')';
    } else if (key === 'yellow') {
      rows = DATA.rows.filter(r => normRiskLocal(r['Risk Level']) === 'YELLOW')
                       .sort((a, b) => (a['Jurisdiction'] || '').localeCompare(b['Jurisdiction'] || ''));
      title = 'YELLOW risk regulations (' + rows.length + ')';
    } else if (key === 'eff90') {
      rows = DATA.metrics.effective_next_90.slice();
      title = 'Effective in next 90 days (' + rows.length + ')';
      opts.showEff = true;
    } else if (key === 'recent') {
      rows = DATA.metrics.recent_changes.slice();
      title = 'Changes in last 30 days (' + rows.length + ')';
      opts.lastReview = true;
    }
    titleEl.textContent = title;
    bodyEl.innerHTML = renderRowsTable(rows, opts);
    panel.classList.add('open');
    panel.scrollIntoView({ behavior:'smooth', block:'nearest' });
  }

  function closeDrill() {
    panel.classList.remove('open');
    document.querySelectorAll('.kpi.clickable').forEach(el => el.classList.remove('active'));
    activeKey = null;
  }

  document.querySelectorAll('.kpi.clickable').forEach(el => {
    el.addEventListener('click', () => openDrill(el.dataset.drill, el));
    el.addEventListener('keydown', e => {
      if (e.key === 'Enter' || e.key === ' ') { e.preventDefault(); openDrill(el.dataset.drill, el); }
    });
  });
  closeEl.addEventListener('click', closeDrill);
})();
</script>
</body>
</html>
"""

def render_html(rows, metrics, latest_updates, state_data, categories, refresh_info):
    payload = {"rows": rows, "metrics": metrics, "latest_updates": latest_updates, "state_data": state_data, "categories": categories, "refresh": refresh_info, "generated": refresh_info["today_iso"]}
    data_json = json.dumps(payload, ensure_ascii=True)
    red = metrics["by_risk"].get("RED", 0)
    yellow = metrics["by_risk"].get("YELLOW", 0)
    total = metrics["total"]
    by_tab_str = " · ".join(f"{n} {t.split()[0]}" for t, n in metrics["by_tab"].items())
    html = HTML_TEMPLATE
    total_regs_state_view = sum(s.get('count', 0) for s in state_data.values())
    today_iso = refresh_info["today_iso"]
    for k, v in {
        "__TODAY__": today_iso, "__TOTAL__": str(total),
        "__TOTAL_REGS__": str(total_regs_state_view),
        "__GENERATED_AT__": refresh_info["generated_at"],
        "__SOURCE_FILE__": refresh_info["source_file"],
        "__SOURCE_MTIME__": refresh_info["source_mtime"],
        "__SOURCE_SIZE_KB__": str(refresh_info["source_size_kb"]),
        "__BY_TAB__": by_tab_str, "__RED__": str(red), "__YELLOW__": str(yellow),
        "__RED_PCT__": str(round(100 * red / total) if total else 0),
        "__YELLOW_PCT__": str(round(100 * yellow / total) if total else 0),
        "__EFF90__": str(metrics["effective_next_90_count"]),
        "__RECENT__": str(metrics["recent_changes_count"]),
    }.items():
        html = html.replace(k, v)
    html = html.replace("__DATA_JSON__", data_json)
    return html

def resolve_freshest_xlsx(requested_path):
    """If sidecar files exist (e.g., _pending_save.xlsx, _NEW.xlsx) that are NEWER than
    the requested file, prefer the newest. Handles the common case where the main xlsx
    is locked open in Excel and scheduled scans write to a sidecar instead."""
    import os
    requested = Path(requested_path)
    if not requested.exists():
        return requested
    # Candidate sidecars in the same directory
    parent = requested.parent
    stem = requested.stem  # 'HR_Compliance_Framework_COMPLETE'
    suffix = requested.suffix  # '.xlsx'
    candidates = [requested]
    for cand_name in [
        f"{stem}_pending_save{suffix}",
        f"{stem}_NEW{suffix}",
        f"{stem}_pending{suffix}",
    ]:
        cand_path = parent / cand_name
        if cand_path.exists():
            candidates.append(cand_path)
    # Filter out backup files (those with _backup or _pre- in name)
    candidates = [c for c in candidates if "backup" not in c.name.lower() and "_pre-" not in c.name.lower()]
    # Pick newest by mtime
    newest = max(candidates, key=lambda c: c.stat().st_mtime)
    if newest != requested:
        print(f"  → Using newer sidecar: {newest.name} ({os.path.getsize(newest):,} bytes, "
              f"mtime {Path(newest).stat().st_mtime:.0f} vs main {requested.stat().st_mtime:.0f})")
    return newest

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="HR_Compliance_Framework_COMPLETE.xlsx")
    parser.add_argument("--out", default="dashboard.html")
    args = parser.parse_args()
    xlsx_path = Path(args.xlsx)
    out_path = Path(args.out)
    if not xlsx_path.is_absolute():
        xlsx_path = Path(__file__).parent / xlsx_path
    if not out_path.is_absolute():
        out_path = Path(__file__).parent / out_path
    xlsx_path = resolve_freshest_xlsx(xlsx_path)
    print(f"Reading: {xlsx_path}")
    law_rows, all_rows, update_tabs = load_all_tab_rows(xlsx_path)
    latest = load_latest_updates(xlsx_path)
    print(f"Law rows: {len(law_rows)} · All-tab rows (incl. surveys): {len(all_rows)} · Update tabs: {len(update_tabs)}")
    print(f"Latest updates: {latest.get('tab_name')} ({len(latest.get('rows', []))} entries)")
    metrics = compute_metrics(law_rows)
    state_data = compute_state_data(all_rows)
    print(f"State data: {len(state_data)} jurisdictions, total regs (deduped): {sum(s['count'] for s in state_data.values())}")
    categories = collect_categories(state_data)
    # Refresh banner data
    from datetime import datetime
    import os
    now_local = datetime.now()
    src_mtime = datetime.fromtimestamp(os.path.getmtime(xlsx_path))
    refresh_info = {
        "generated_at": now_local.strftime("%Y-%m-%d %I:%M:%S %p"),
        "generated_at_short": now_local.strftime("%Y-%m-%d %H:%M"),
        "source_file": xlsx_path.name,
        "source_mtime": src_mtime.strftime("%Y-%m-%d %I:%M %p"),
        "source_size_kb": round(os.path.getsize(xlsx_path) / 1024),
        "today_iso": date.today().isoformat(),
    }
    html = render_html(law_rows, metrics, latest, state_data, categories, refresh_info)
    out_path.write_text(html, encoding="utf-8")
    print(f"Wrote: {out_path}  ({out_path.stat().st_size:,} bytes) @ {refresh_info['generated_at']}")

if __name__ == "__main__":
    main()
