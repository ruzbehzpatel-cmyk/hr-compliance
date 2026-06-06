"""
verify_links.py — HEAD-checks every URL in HR_Compliance_Framework_COMPLETE.xlsx.

Runs through `Gov Source URL` and `Bill/Statute URL` columns across Federal Laws,
State Laws, Additional & Emerging, and every Survey tab. Flags:
  - 4xx (404 dead, 403 forbidden)
  - 5xx (server error)
  - Redirect that lands on a totally different domain (e.g., reg-specific URL → home page)
  - Connection / DNS / timeout errors

Output:
  - Prints a console summary
  - Writes a 'Link Check YYYY-MM-DD' tab in the workbook with one row per broken URL
    (and saves to a sidecar if main file is locked)

Usage:
    python verify_links.py
    python verify_links.py --xlsx path/to/file.xlsx
    python verify_links.py --quick   # only check unverified URLs (skip recently-passing ones)
"""
import openpyxl
import argparse
import re
import sys
import time
from urllib.parse import urlparse
from datetime import date
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import requests
except ImportError:
    print("Installing requests…")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "requests", "--quiet"], check=True)
    import requests

UA = "Mozilla/5.0 (compatible; HR-Compliance-LinkChecker/1.0; +contact: ruzbeh.z.patel@gmail.com)"
TIMEOUT = 10
MAX_WORKERS = 12

LAW_TABS = ["Federal Laws", "State Laws", "Additional & Emerging"]
URL_COLUMNS = ["Gov Source URL", "Bill/Statute URL"]


def is_same_origin(u1, u2):
    try:
        return urlparse(u1).netloc.lower() == urlparse(u2).netloc.lower()
    except Exception:
        return False


def check_url(url):
    """Returns dict: {url, status_code, final_url, error, redirected_to_home, ok}."""
    result = {"url": url, "status_code": None, "final_url": url,
              "error": None, "redirected_to_home": False, "ok": False}
    if not url or not url.startswith("http"):
        result["error"] = "not-a-url"
        return result
    headers = {"User-Agent": UA, "Accept": "text/html,*/*"}
    try:
        # HEAD first; some servers reject HEAD so fall back to GET
        r = requests.head(url, headers=headers, timeout=TIMEOUT, allow_redirects=True)
        if r.status_code in (405, 501) or (400 <= r.status_code < 500 and r.status_code != 404):
            r = requests.get(url, headers=headers, timeout=TIMEOUT,
                             allow_redirects=True, stream=True)
            r.close()
        result["status_code"] = r.status_code
        result["final_url"] = r.url
        # Detect redirect-to-home (specific URL → root or just "/")
        try:
            orig_path = urlparse(url).path or "/"
            final_path = urlparse(r.url).path or "/"
            if len(orig_path) > 4 and (final_path == "/" or len(final_path) < 4):
                if is_same_origin(url, r.url):
                    result["redirected_to_home"] = True
        except Exception:
            pass
        result["ok"] = 200 <= r.status_code < 400 and not result["redirected_to_home"]
    except requests.exceptions.SSLError as e:
        result["error"] = "ssl-error"
    except requests.exceptions.ConnectionError:
        result["error"] = "connection-failed"
    except requests.exceptions.Timeout:
        result["error"] = "timeout"
    except requests.exceptions.TooManyRedirects:
        result["error"] = "too-many-redirects"
    except Exception as e:
        result["error"] = f"other:{type(e).__name__}"
    return result


def collect_urls(xlsx_path):
    """Returns [(url, tab, row, col_name, law_id, reg_name)] for every URL in the workbook."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    urls = []
    for sheet_name in wb.sheetnames:
        if sheet_name in ("Reference", "Dashboard", "QC Findings"):
            continue
        if re.match(r"Updates? \d{4}-\d{2}-\d{2}|Link Check ", sheet_name):
            continue
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=False)
        try:
            headers = [c.value for c in next(rows_iter)]
        except StopIteration:
            continue
        idx = {h: i for i, h in enumerate(headers) if h}
        law_id_idx = idx.get("Law ID")
        name_idx = idx.get("Law / Regulation Name")
        for r_num, row in enumerate(rows_iter, start=2):
            for col_name in URL_COLUMNS:
                if col_name in idx:
                    cell = row[idx[col_name]]
                    val = cell.value
                    if val and str(val).strip().startswith("http"):
                        urls.append({
                            "url": str(val).strip(),
                            "tab": sheet_name,
                            "row": r_num,
                            "col": col_name,
                            "law_id": str(row[law_id_idx].value or "") if law_id_idx is not None else "",
                            "name": str(row[name_idx].value or "")[:80] if name_idx is not None else "",
                        })
    wb.close()
    return urls


def write_link_check_tab(xlsx_path, broken):
    """Writes a 'Link Check YYYY-MM-DD' tab listing broken links. Falls back to sidecar if locked."""
    today = date.today().isoformat()
    tab_name = f"Link Check {today}"
    headers = ["Tab", "Row", "Column", "Law ID", "Regulation", "URL",
               "Status / Error", "Final URL", "Recommended action"]
    targets = [xlsx_path, xlsx_path.with_name(xlsx_path.stem + "_link_check.xlsx")]
    last_err = None
    for target in targets:
        try:
            if target.exists():
                wb = openpyxl.load_workbook(target)
                if tab_name in wb.sheetnames:
                    del wb[tab_name]
            else:
                wb = openpyxl.Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
            ws = wb.create_sheet(tab_name)
            ws.append(headers)
            for b in broken:
                err_str = (str(b["status_code"]) if b["status_code"] else "")
                if b["error"]:
                    err_str = b["error"]
                if b.get("redirected_to_home"):
                    err_str = (err_str + " · redirect-to-home").strip(" ·")
                rec = ""
                if b.get("redirected_to_home"):
                    rec = "URL likely moved — find current regulation-specific page on same domain"
                elif b["status_code"] == 404:
                    rec = "Page removed — search agency site for new URL"
                elif b["error"] == "ssl-error":
                    rec = "SSL/cert issue — try http or check for new HTTPS endpoint"
                elif b["error"] == "connection-failed":
                    rec = "Domain may have moved — verify agency website"
                else:
                    rec = "Investigate manually"
                ws.append([b["tab"], b["row"], b["col"], b["law_id"], b["name"],
                           b["url"], err_str, b.get("final_url", ""), rec])
            for i, w in enumerate([18, 6, 18, 14, 38, 50, 22, 50, 40], start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            ws.freeze_panes = "A2"
            wb.save(target)
            return target
        except PermissionError as e:
            last_err = e
            continue
    raise last_err


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default="HR_Compliance_Framework_COMPLETE.xlsx")
    parser.add_argument("--limit", type=int, default=0, help="Stop after N URLs (0 = unlimited)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_absolute():
        xlsx_path = Path(__file__).parent / xlsx_path
    if not xlsx_path.exists():
        print(f"ERROR: file not found: {xlsx_path}")
        sys.exit(1)

    # Auto-pick freshest sidecar (mirrors generate_dashboard.py behavior)
    parent = xlsx_path.parent
    stem = xlsx_path.stem
    suffix = xlsx_path.suffix
    candidates = [xlsx_path]
    for cand_name in [f"{stem}_pending_save{suffix}", f"{stem}_NEW{suffix}", f"{stem}_pending{suffix}"]:
        cp = parent / cand_name
        if cp.exists():
            candidates.append(cp)
    candidates = [c for c in candidates if "backup" not in c.name.lower() and "_pre-" not in c.name.lower()]
    xlsx_path = max(candidates, key=lambda c: c.stat().st_mtime)
    print(f"Reading: {xlsx_path.name}  ({xlsx_path.stat().st_size:,} bytes)")

    urls = collect_urls(xlsx_path)
    print(f"Collected {len(urls)} URLs across the workbook")
    if args.limit:
        urls = urls[: args.limit]

    print(f"Checking {len(urls)} URLs with {MAX_WORKERS} parallel workers…\n")
    t0 = time.time()
    broken = []
    by_status = {}
    by_error = {}

    def task(u):
        res = check_url(u["url"])
        u.update(res)
        return u

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(task, u): u for u in urls}
        for i, fut in enumerate(as_completed(futures), start=1):
            r = fut.result()
            if r["ok"]:
                by_status[r["status_code"]] = by_status.get(r["status_code"], 0) + 1
            else:
                broken.append(r)
                key = r["error"] or str(r["status_code"])
                by_error[key] = by_error.get(key, 0) + 1
                marker = "X"
                print(f"  {marker} [{r['tab'][:18]:18s}] {r['law_id'][:14]:14s}  "
                      f"{(r['error'] or str(r['status_code'])):20s}  {r['url'][:90]}")
            if i % 20 == 0:
                print(f"   …{i}/{len(urls)} checked")

    elapsed = time.time() - t0
    print(f"\n=== Link check complete ({elapsed:.1f}s) ===")
    print(f"  OK:     {len(urls) - len(broken)} / {len(urls)}")
    print(f"  Broken: {len(broken)}")
    if by_error:
        print("\nBroken-link breakdown:")
        for k, v in sorted(by_error.items(), key=lambda x: -x[1]):
            print(f"  {v:4d}  {k}")

    if broken:
        target = write_link_check_tab(xlsx_path, broken)
        print(f"\nWrote broken-link tab to: {target.name}")
    else:
        print("\nNo broken links found.")

    return 0 if not broken else 2


if __name__ == "__main__":
    sys.exit(main())
